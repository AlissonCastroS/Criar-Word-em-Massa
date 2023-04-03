import os
import openpyxl
from docxtpl import DocxTemplate
import PySimpleGUI as sg
from docx2pdf import convert

# abertura de janela
class TelaPy:
    def __init__(self):
        font = ('Helvetica', 11)
        sg.theme('DarkTeal4')
        sg.set_options(font=font)

        # botões e campos
        layout = [
            [sg.Text('Selecionar planilha de dados', size=(25, 1)), sg.Input(), sg.FileBrowse('Selecionar...')],
            [sg.Text('Selecionar pasta de destino', size=(25, 1)), sg.Input(), sg.FolderBrowse('Selecionar...')],
            [sg.Text('Data do documento', size=(25, 1)), sg.Input()],
            [sg.Text(''), sg.Input(key='valor', visible=False)],
            [sg.Submit('Iniciar')],
            [sg.Submit('Iniciar e converter'), sg.Cancel('Cancelar')]
        ]

        # Janela
        self.window = sg.Window('titulo de janela', layout, icon='meu_icon.ico')

    def run(self):
        while True:
            event, values = self.window.read()
            if event == 'Enviar':
                valor = values['valor']
                print(f'Valor inserido: {valor}')
            elif event == 'Cancelar' or event == sg.WIN_CLOSED:
                break

            print(f'{event} iniciado')

            # escolhendo o excel como banco de dados
            workbook = openpyxl.load_workbook(f"{values[0]}")
            sheet = workbook.active

            output_folder = values[1]

            # modelo a ser formatado
            doc = DocxTemplate("C:\\caminho_do_arquivo_word\modelo.docx")

            # adcionar contexto com colunas a serem lidas e as linhas que serão incluidas
            for row in sheet.iter_rows(values_only=True):
                context = {
                    # campo da data ou texto desejado
                    "campo de escrita": f"{values[2]}",
                    # colunas a serem usadas
                    "coluna1": row[0],
                    "coluna2": row[1]
                }

                # Renderizar o modelo com o 'context'
                doc.render(context)

                # Salvar o documento com um nome baseado na linha atual
                filename = os.path.join(output_folder, f"{row[0]}.docx")
                doc.save(filename)

            if event == 'Iniciar e converter':
                for row in sheet.iter_rows(values_only=True):
                    # Converter o arquivo .docx para PDF
                    pdf_filename = os.path.join(output_folder, f"{row[0]}.pdf")
                    convert(os.path.join(output_folder, f"{row[0]}.docx"), pdf_filename)

        self.window.close()

# fechamento da tela
if __name__ == '__main__':
    TelaPy().run()